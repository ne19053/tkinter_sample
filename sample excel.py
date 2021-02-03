'''
tkinter application
input data and save to excel file
'''

import tkinter as tk
import openpyxl

WORKBOOK = None

if WORKBOOK is None:
    wb = openpyxl.Workbook()
    WORKBOOK = 'Sample.xls'
else:
    wb = openpyxl.load_workbook(WORKBOOK)

ws = wb.worksheets[0]
c_row = 2

root = tk.Tk()
def update_ws():
    global c_row
    col_id = ws.cell(c_row,1)
    col_name = ws.cell(c_row,2)
    col_result = ws.cell(c_row,3)
    col_id.value = en_id.get()
    col_name.value = en_name.get()
    col_result.value = en_result.get()
    c_row += 1

def commit_change():
    wb.save(WORKBOOK)

lb_id = tk.Label(text='Student ID')
en_id = tk.Entry()
lb_name = tk.Label(text='Name')
en_name = tk.Entry
lb_result = tk.Label(text='Result')
en_result = tk.Entry
bt1 = tk.Button(text='Enter data',command=update_ws)
bt2 = tk.Button(text='Commit Change',command=commit_change)
[w.pack() for w in (lb_id, en_id, lb_name, en_name, lb_result, en_result, bt1, bt2)]

root.mainloop()

